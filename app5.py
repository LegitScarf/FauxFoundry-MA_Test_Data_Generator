import streamlit as st
import json
import pandas as pd
import io
import re
from typing import Dict, Any, List
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="FauxFoundry v2",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Orbitron:wght@400;700;900&family=Share+Tech+Mono&display=swap');

  html, body, [class*="css"] {
      background-color: #0a0a0f;
      color: #c8d8e8;
      font-family: 'Share Tech Mono', monospace;
  }
  .stApp { background: linear-gradient(135deg, #0a0a0f 0%, #0d1117 50%, #0a0f1a 100%); }

  h1, h2, h3 { font-family: 'Orbitron', monospace !important; color: #00d4ff !important; letter-spacing: 0.15em; }

  .stButton > button {
      background: linear-gradient(90deg, #003366, #0055aa);
      color: #00d4ff;
      border: 1px solid #00d4ff;
      border-radius: 4px;
      font-family: 'Orbitron', monospace;
      font-size: 0.75rem;
      letter-spacing: 0.1em;
      transition: all 0.3s;
  }
  .stButton > button:hover { background: linear-gradient(90deg, #0055aa, #0077cc); box-shadow: 0 0 12px #00d4ff55; }
  .stButton > button:disabled { opacity: 0.35; cursor: not-allowed; }

  .stTextArea textarea, .stTextInput input {
      background: #0d1117 !important;
      border: 1px solid #1e3a5f !important;
      color: #c8d8e8 !important;
      font-family: 'Share Tech Mono', monospace !important;
      border-radius: 4px;
  }
  .stTextArea textarea:focus, .stTextInput input:focus { border-color: #00d4ff !important; box-shadow: 0 0 8px #00d4ff33 !important; }

  .stDownloadButton > button {
      background: linear-gradient(90deg, #003322, #006644) !important;
      color: #00ff88 !important;
      border: 1px solid #00ff88 !important;
      font-family: 'Orbitron', monospace;
      font-size: 0.7rem;
      letter-spacing: 0.08em;
      width: 100%;
  }
  .stDownloadButton > button:hover { box-shadow: 0 0 12px #00ff8855 !important; }

  .stProgress > div > div { background: linear-gradient(90deg, #0055aa, #00d4ff) !important; }

  .stMetric { background: #0d1117; border: 1px solid #1e3a5f; border-radius: 6px; padding: 1rem; }
  .stMetric label { color: #7090b0 !important; font-size: 0.7rem !important; letter-spacing: 0.1em; }
  .stMetric [data-testid="stMetricValue"] { color: #00d4ff !important; font-family: 'Orbitron', monospace; }

  .stTabs [data-baseweb="tab-list"] { background: #0d1117; border-bottom: 1px solid #1e3a5f; gap: 4px; }
  .stTabs [data-baseweb="tab"] { background: #111827; color: #7090b0; border: 1px solid #1e3a5f; border-radius: 4px 4px 0 0; font-family: 'Orbitron', monospace; font-size: 0.65rem; letter-spacing: 0.08em; }
  .stTabs [aria-selected="true"] { background: #0d2244 !important; color: #00d4ff !important; border-color: #00d4ff !important; }

  .hero { text-align: center; padding: 2.5rem 0 1.5rem; }
  .hero h1 { font-size: 3.2rem; letter-spacing: 0.4em; background: linear-gradient(90deg, #00d4ff, #0088ff, #00d4ff); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
  .hero p { color: #4080b0; letter-spacing: 0.25em; font-size: 0.85rem; margin-top: -0.5rem; }

  .status-ok  { color: #00ff88; font-family: 'Share Tech Mono', monospace; }
  .status-off { color: #ffaa00; font-family: 'Share Tech Mono', monospace; }
  .status-err { color: #ff4444; font-family: 'Share Tech Mono', monospace; }

  .panel { background: #0d1117; border: 1px solid #1e3a5f; border-radius: 6px; padding: 1rem 1.25rem; margin-bottom: 1rem; }
  .agent-label { color: #0088ff; font-family: 'Orbitron', monospace; font-size: 0.7rem; letter-spacing: 0.1em; margin-bottom: 0.4rem; }

  div[data-testid="stSidebar"] { background: #0a0d14; border-right: 1px solid #1e3a5f; }
  .stSelectbox > div { background: #0d1117 !important; border: 1px solid #1e3a5f !important; color: #c8d8e8 !important; }

  .rel-badge { display: inline-block; background: #0d2244; border: 1px solid #0055aa; border-radius: 3px; padding: 2px 8px; margin: 2px; font-size: 0.7rem; color: #00d4ff; font-family: 'Share Tech Mono', monospace; }
  footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ── Generator class ────────────────────────────────────────────────────────────
class SyntheticDataGenerator:
    def __init__(self):
        self.client: OpenAI | None = None

    def setup(self, api_key: str) -> bool:
        try:
            self.client = OpenAI(api_key=api_key)
            # quick connectivity check
            self.client.models.list()
            return True
        except Exception as e:
            st.error(f"OpenAI setup failed: {e}")
            return False

    # ── Agent 1 ── schema ──────────────────────────────────────────────────────
    def generate_schema(self, brief: str, mode: str, num_rows: int) -> Dict[str, Any]:
        if not self.client:
            raise RuntimeError("OpenAI client not configured")

        if mode == "Simple (Single Table)":
            system = "You are a data schema designer. Return ONLY valid JSON, no markdown fences."
            user = f"""
From the user request below, build a single-table schema JSON:
{{
  "tables": [
    {{
      "name": "<TableName>",
      "primary_key": "<pk_column>",
      "foreign_keys": [],
      "columns": {{ "<col>": "<type>", ... }}
    }}
  ],
  "num_rows": {num_rows},
  "dataset_description": "<short description>"
}}
Supported types: string, integer, float, boolean, date, email, phone, address, url
Request: {brief}
"""
        else:  # Relational
            system = "You are a relational database schema designer. Return ONLY valid JSON, no markdown fences."
            user = f"""
Analyse the project brief and produce a multi-table relational schema JSON.
Rules:
- Identify ALL tables mentioned (or implied) in the brief.
- Identify all primary keys and foreign key relationships.
- Order tables so parent tables (no foreign keys) come first.
- Use {num_rows} as the base row count for the root/parent table; child tables get {num_rows} rows too unless the brief specifies otherwise.
- Supported column types: string, integer, float, boolean, date, email, phone, address, url

Output EXACTLY this structure:
{{
  "tables": [
    {{
      "name": "<TableName>",
      "primary_key": "<pk_column>",
      "foreign_keys": [
        {{ "column": "<fk_col>", "references_table": "<ParentTable>", "references_column": "<pk_col>" }}
      ],
      "columns": {{ "<col>": "<type>", ... }}
    }}
  ],
  "num_rows": {num_rows},
  "dataset_description": "<short description>"
}}

Project brief:
{brief}
"""
        resp = self.client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
            temperature=0.2,
            max_tokens=1200,
        )
        raw = resp.choices[0].message.content.strip()
        raw = self._strip_fences(raw)
        return json.loads(raw)

    # ── Agent 2 ── data rows ───────────────────────────────────────────────────
    def generate_table_data(
        self,
        table: Dict[str, Any],
        num_rows: int,
        parent_ids: Dict[str, List[Any]] | None = None,
    ) -> pd.DataFrame:
        if not self.client:
            raise RuntimeError("OpenAI client not configured")

        fk_context = ""
        if parent_ids:
            lines = []
            for fk in table.get("foreign_keys", []):
                col = fk["column"]
                ref_table = fk["references_table"]
                ref_col = fk["references_column"]
                ids = parent_ids.get(f"{ref_table}.{ref_col}", [])
                # Show first 60 IDs so the prompt doesn't blow up for large sets
                sample = ids[:60]
                lines.append(
                    f'- Column "{col}" must contain only values from this list '
                    f'(the existing {ref_table}.{ref_col} values): {sample}'
                )
            fk_context = "\nForeign-key constraints (STRICTLY enforce):\n" + "\n".join(lines)

        prompt = f"""
Generate a realistic synthetic dataset for the table "{table['name']}".

Columns and types:
{json.dumps(table['columns'], indent=2)}

Primary key column: {table.get('primary_key', 'N/A')} — values must be unique.
Number of rows: {num_rows}
{fk_context}

Requirements:
1. Produce realistic, diverse, internally-consistent data.
2. Primary key values must be unique across all rows.
3. Honour every foreign-key constraint exactly — use ONLY the provided IDs.
4. Return a markdown table (pipe-delimited) with a header row. Nothing else.
"""
        resp = self.client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a synthetic data generator. Output only a markdown table."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.75,
            max_tokens=3500,
        )
        md = resp.choices[0].message.content
        return self._markdown_to_df(md)

    # ── Agent 3 ── validation + repair ────────────────────────────────────────
    def validate_and_repair(
        self,
        tables: Dict[str, pd.DataFrame],
        schema: Dict[str, Any],
    ) -> Dict[str, pd.DataFrame]:
        """Enforce FK integrity in-process (no LLM call needed)."""
        repaired = dict(tables)
        for tbl in schema["tables"]:
            name = tbl["name"]
            if name not in repaired:
                continue
            df = repaired[name].copy()
            for fk in tbl.get("foreign_keys", []):
                col = fk["column"]
                ref_t = fk["references_table"]
                ref_c = fk["references_column"]
                if ref_t not in repaired or col not in df.columns:
                    continue
                valid_ids = repaired[ref_t][ref_c].tolist()
                # Replace invalid FK values by cycling through valid ones
                mask = ~df[col].astype(str).isin([str(v) for v in valid_ids])
                if mask.any():
                    cycle = [valid_ids[i % len(valid_ids)] for i in range(mask.sum())]
                    df.loc[mask, col] = cycle
            repaired[name] = df
        return repaired

    # ── helpers ────────────────────────────────────────────────────────────────
    @staticmethod
    def _strip_fences(text: str) -> str:
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            text = text.split("```")[1].split("```")[0]
        return text.strip()

    @staticmethod
    def _markdown_to_df(md: str) -> pd.DataFrame:
        lines = [l for l in md.strip().split("\n") if "|" in l]
        # remove separator rows
        lines = [l for l in lines if not re.match(r"^[\|\-\:\s]+$", l)]
        if len(lines) < 2:
            raise ValueError("Model returned no usable table rows.")
        rows = []
        for l in lines:
            cells = [c.strip() for c in l.split("|")]
            cells = [c for c in cells if c != ""]
            rows.append(cells)
        headers = rows[0]
        data = rows[1:]
        # Pad / trim so every row matches header length
        n = len(headers)
        data = [r[:n] + [""] * (n - len(r)) for r in data]
        return pd.DataFrame(data, columns=headers)

    # ── Excel builder ──────────────────────────────────────────────────────────
    @staticmethod
    def build_excel(tables: Dict[str, pd.DataFrame], schema: Dict[str, Any]) -> bytes:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        HDR_FILL   = PatternFill("solid", fgColor="0D2244")
        HDR_FONT   = Font(name="Calibri", bold=True, color="00D4FF", size=11)
        CELL_FONT  = Font(name="Calibri", size=10, color="C8D8E8")
        ALT_FILL   = PatternFill("solid", fgColor="0A1628")
        BORDER_SIDE = Side(style="thin", color="1E3A5F")
        CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

        for tbl_def in schema["tables"]:
            name = tbl_def["name"]
            if name not in tables:
                continue
            df = tables[name]
            ws = wb.create_sheet(title=name[:31])  # sheet name max 31 chars

            ws.sheet_view.showGridLines = False
            ws.sheet_properties.tabColor = "00D4FF"

            # Header row
            for ci, col in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = CELL_BORDER

            # Data rows
            for ri, row in df.iterrows():
                excel_row = ri + 2
                fill = ALT_FILL if ri % 2 == 1 else None
                for ci, val in enumerate(row, start=1):
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.font = CELL_FONT
                    cell.border = CELL_BORDER
                    cell.alignment = Alignment(vertical="center")
                    if fill:
                        cell.fill = fill

            # Auto column width
            for ci, col in enumerate(df.columns, start=1):
                max_len = max(
                    len(str(col)),
                    *[len(str(v)) for v in df.iloc[:, ci - 1]],
                )
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"

        # ── Relationships info sheet ──────────────────────────────────────────
        ws_rel = wb.create_sheet(title="📊 Schema Info")
        ws_rel.sheet_properties.tabColor = "FF8800"
        ws_rel.sheet_view.showGridLines = False

        ws_rel["A1"] = "FauxFoundry v2 — Schema & Relationships"
        ws_rel["A1"].font = Font(name="Calibri", bold=True, color="00D4FF", size=14)
        ws_rel["A1"].fill = PatternFill("solid", fgColor="0D1117")

        ws_rel["A3"] = "Dataset Description"
        ws_rel["A3"].font = Font(bold=True, color="7090B0")
        ws_rel["B3"] = schema.get("dataset_description", "")
        ws_rel["B3"].font = Font(color="C8D8E8")

        ws_rel["A5"] = "Table"
        ws_rel["B5"] = "Primary Key"
        ws_rel["C5"] = "Foreign Keys"
        ws_rel["D5"] = "Row Count"
        for col in "ABCD":
            c = ws_rel[f"{col}5"]
            c.font = Font(bold=True, color="00D4FF")
            c.fill = PatternFill("solid", fgColor="0D2244")
            c.border = CELL_BORDER

        for ri, tbl_def in enumerate(schema["tables"], start=6):
            name = tbl_def["name"]
            pk   = tbl_def.get("primary_key", "")
            fks  = "; ".join(
                f"{fk['column']} → {fk['references_table']}.{fk['references_column']}"
                for fk in tbl_def.get("foreign_keys", [])
            ) or "—"
            row_count = len(tables.get(name, pd.DataFrame()))
            for ci, val in enumerate([name, pk, fks, row_count], start=1):
                cell = ws_rel.cell(row=ri, column=ci, value=val)
                cell.font = Font(name="Calibri", size=10, color="C8D8E8")
                cell.border = CELL_BORDER
                if ri % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="0A1628")

        for col, width in zip("ABCD", [22, 18, 48, 12]):
            ws_rel.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# ── Streamlit UI ───────────────────────────────────────────────────────────────
def main():
    # Hero
    st.markdown("""
    <div class="hero">
      <h1>FAUXFOUNDRY</h1>
      <p>RELATIONAL SYNTHETIC DATA GENERATOR &nbsp;·&nbsp; V2.0</p>
    </div>
    """, unsafe_allow_html=True)

    if "generator" not in st.session_state:
        st.session_state.generator = SyntheticDataGenerator()
    if "schema" not in st.session_state:
        st.session_state.schema = None
    if "tables" not in st.session_state:
        st.session_state.tables = {}

    gen: SyntheticDataGenerator = st.session_state.generator

    # ── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("## ⚡ NEURAL INTERFACE")

        api_key = st.text_input("🧠 OpenAI API Key", type="password")
        if api_key and st.button("🔧 INITIALIZE"):
            with st.spinner("Connecting…"):
                ok = gen.setup(api_key)
                st.session_state.api_ok = ok
            if ok:
                st.markdown('<p class="status-ok">⚡ Neural link active</p>', unsafe_allow_html=True)
            else:
                st.markdown('<p class="status-err">❌ Connection failed</p>', unsafe_allow_html=True)

        linked = st.session_state.get("api_ok", False)
        if linked:
            st.markdown('<p class="status-ok">🟢 NEURAL LINK ACTIVE</p>', unsafe_allow_html=True)
        else:
            st.markdown('<p class="status-off">🟡 NEURAL LINK OFFLINE</p>', unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### ⚙️ SYNTHESIS PARAMETERS")

        mode = st.selectbox(
            "Mode",
            ["Simple (Single Table)", "Advanced (Relational / Multi-Table)"],
            help="Choose Simple for a single table, Advanced for multi-table relational datasets",
        )
        num_rows = st.slider("Base row count", min_value=10, max_value=500, value=50, step=10)

        st.markdown("---")
        st.markdown("### 📖 QUICK GUIDE")
        if mode == "Simple (Single Table)":
            st.caption("Describe your table in plain English. E.g.: *Create a dataset of 50 employees with name, age, department, salary, and hire date.*")
        else:
            st.caption("Paste your full project brief or table requirements. The engine will extract all tables, columns, and relationships automatically.")

    # ── Main layout ────────────────────────────────────────────────────────────
    left, right = st.columns([2, 1])

    with left:
        st.markdown("## 📡 SYNTHESIS REQUEST")

        placeholder = (
            "Describe your single table here.\n\nExample: Create a dataset about employees with columns for name, age, department, salary, and hire date."
            if mode == "Simple (Single Table)"
            else
            "Paste your full project brief here.\n\nExample:\n\nCustomer Churn Analysis\nTable: Customers — CustomerID, Name, Age, Gender, Region, Tenure, Churn\nTable: Usage — CustomerID, CallMinutes, DataUsage, MessagesSent\nTable: Feedback — CustomerID, SatisfactionScore, Complaints\nRelationships: Usage.CustomerID → Customers.CustomerID; Feedback.CustomerID → Customers.CustomerID"
        )
        brief = st.text_area("Transmit your requirements:", placeholder=placeholder, height=180)

        ready = linked and bool(brief.strip())

        if st.button("⚡ INITIATE DATA SYNTHESIS", disabled=not ready):
            st.session_state.schema  = None
            st.session_state.tables  = {}

            prog = st.progress(0)
            status = st.empty()

            try:
                # ── Agent 1 ── schema ──────────────────────────────────────────
                status.markdown('<p class="agent-label">🤖 AGENT 01 — Parsing schema…</p>', unsafe_allow_html=True)
                prog.progress(5)
                with st.spinner("Generating schema…"):
                    schema = gen.generate_schema(brief, mode, num_rows)
                st.session_state.schema = schema
                prog.progress(20)

                st.markdown('<div class="panel"><div class="agent-label">AGENT 01 OUTPUT — SCHEMA MATRIX</div></div>', unsafe_allow_html=True)
                st.json(schema)

                tables_def: List[Dict] = schema["tables"]
                generated: Dict[str, pd.DataFrame] = {}
                parent_ids: Dict[str, List] = {}

                total = len(tables_def)
                for idx, tbl in enumerate(tables_def):
                    tname = tbl["name"]
                    pct_start = 20 + int(60 * idx / total)
                    pct_end   = 20 + int(60 * (idx + 1) / total)

                    status.markdown(
                        f'<p class="agent-label">🤖 AGENT 02 — Generating table {idx+1}/{total}: <strong>{tname}</strong>…</p>',
                        unsafe_allow_html=True,
                    )
                    prog.progress(pct_start)

                    with st.spinner(f"Fabricating {tname}…"):
                        df = gen.generate_table_data(tbl, num_rows, parent_ids if tbl.get("foreign_keys") else None)
                    generated[tname] = df

                    # Register this table's PK values for child FK constraints
                    pk = tbl.get("primary_key")
                    if pk and pk in df.columns:
                        parent_ids[f"{tname}.{pk}"] = df[pk].tolist()

                    prog.progress(pct_end)

                # ── Agent 3 ── validate & repair ───────────────────────────────
                status.markdown('<p class="agent-label">🤖 AGENT 03 — Validating referential integrity…</p>', unsafe_allow_html=True)
                prog.progress(85)
                with st.spinner("Running integrity checks…"):
                    generated = gen.validate_and_repair(generated, schema)

                st.session_state.tables = generated
                prog.progress(100)
                status.markdown('<p class="status-ok">✅ DATA SYNTHESIS COMPLETE</p>', unsafe_allow_html=True)
                st.markdown('<p class="status-ok">⚡ All tables generated and verified.</p>', unsafe_allow_html=True)

            except Exception as e:
                st.markdown(f'<p class="status-err">❌ SYNTHESIS ERROR: {e}</p>', unsafe_allow_html=True)
                prog.empty()
                status.empty()

        # ── Table preview tabs ─────────────────────────────────────────────────
        if st.session_state.tables:
            st.markdown("## 📊 GENERATED DATA MATRICES")
            tab_names = list(st.session_state.tables.keys())
            tabs = st.tabs([f"📋 {n}" for n in tab_names])
            for tab, name in zip(tabs, tab_names):
                with tab:
                    df = st.session_state.tables[name]
                    st.dataframe(df, use_container_width=True, height=300)

            # Relationship summary
            if st.session_state.schema:
                rel_lines = []
                for tbl in st.session_state.schema["tables"]:
                    for fk in tbl.get("foreign_keys", []):
                        rel_lines.append(
                            f'<span class="rel-badge">{tbl["name"]}.{fk["column"]} → {fk["references_table"]}.{fk["references_column"]}</span>'
                        )
                if rel_lines:
                    st.markdown("**🔗 Active Relationships:**", unsafe_allow_html=False)
                    st.markdown(" ".join(rel_lines), unsafe_allow_html=True)

    # ── Right panel — metrics + download ──────────────────────────────────────
    with right:
        st.markdown("## 📈 SYNTHESIS METRICS")

        if st.session_state.schema and st.session_state.tables:
            schema  = st.session_state.schema
            tables  = st.session_state.tables

            st.metric("Tables Generated", len(tables))
            total_cols = sum(len(df.columns) for df in tables.values())
            st.metric("Total Columns",    total_cols)
            total_rows = sum(len(df) for df in tables.values())
            st.metric("Total Rows",       total_rows)

            st.markdown("---")
            st.markdown("### TABLE BREAKDOWN")
            for name, df in tables.items():
                st.markdown(
                    f'<div class="panel"><span style="color:#00d4ff;font-size:0.75rem;">'
                    f'▸ {name}</span><br/>'
                    f'<span style="color:#4080b0;font-size:0.7rem;">{len(df.columns)} cols &nbsp;·&nbsp; {len(df)} rows</span></div>',
                    unsafe_allow_html=True,
                )

        st.markdown("---")
        st.markdown("## 💾 DATA EXTRACTION")

        if st.session_state.tables and st.session_state.schema:
            # Excel multi-sheet
            with st.spinner("Building Excel workbook…"):
                excel_bytes = gen.build_excel(st.session_state.tables, st.session_state.schema)

            st.download_button(
                label="📥 DOWNLOAD EXCEL (.xlsx)",
                data=excel_bytes,
                file_name="fauxfoundry_dataset.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # Individual CSV per table
            st.markdown("**Per-table CSV downloads:**")
            for name, df in st.session_state.tables.items():
                csv_bytes = df.to_csv(index=False).encode()
                st.download_button(
                    label=f"📥 {name}.csv",
                    data=csv_bytes,
                    file_name=f"fauxfoundry_{name.lower()}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key=f"csv_{name}",
                )

            # Schema JSON
            st.download_button(
                label="📥 SCHEMA JSON",
                data=json.dumps(st.session_state.schema, indent=2),
                file_name="fauxfoundry_schema.json",
                mime="application/json",
                use_container_width=True,
            )
        else:
            st.info("🔄 Run synthesis to unlock downloads.")

    # ── Footer ─────────────────────────────────────────────────────────────────
    st.markdown("""
    <hr style="margin-top:3rem;opacity:0.15;"/>
    <div style="text-align:center;opacity:0.4;padding:1rem;font-size:0.7rem;letter-spacing:0.15em;">
      FAUXFOUNDRY V2.0 &nbsp;·&nbsp; RELATIONAL SYNTHETIC DATA ENGINE<br/>
      POWERED BY GPT-4o-mini &nbsp;·&nbsp; BUILT WITH STREAMLIT<br/>
      ⚡ SYNTHETIC INTELLIGENCE LABORATORY ⚡
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
